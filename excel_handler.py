"""
Manipulador inteligente de planilhas e arquivos de telefone.
Detecta automaticamente numeros de telefone em QUALQUER coluna, linha ou formato.
Suporta Excel (.xlsx/.xls), CSV e TXT.
"""

import os
import re
import csv
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Cores para os resultados
CORES = {
    "PESSOA_ATENDEU":     "92D050",  # Verde
    "CAIXA_POSTAL":       "FFC000",  # Amarelo
    "FORA_DE_AREA":       "FF6600",  # Laranja
    "NUMERO_INEXISTENTE": "FF0000",  # Vermelho
    "BLOQUEADO":          "A020F0",  # Roxo
    "OCUPADO":            "00B0F0",  # Azul
    "NAO_ATENDEU":        "BFBFBF",  # Cinza
    "ERRO":               "808080",  # Cinza escuro
}

# DDDs validos do Brasil (2 digitos)
DDDS_BRASIL = {
    11, 12, 13, 14, 15, 16, 17, 18, 19,  # SP
    21, 22, 24,                            # RJ
    27, 28,                                # ES
    31, 32, 33, 34, 35, 37, 38,           # MG
    41, 42, 43, 44, 45, 46,               # PR
    47, 48, 49,                            # SC
    51, 53, 54, 55,                        # RS
    61,                                    # DF
    62, 64,                                # GO
    63,                                    # TO
    65, 66,                                # MT
    67,                                    # MS
    68,                                    # AC
    69,                                    # RO
    71, 73, 74, 75, 77,                   # BA
    79,                                    # SE
    81, 87,                                # PE
    82,                                    # AL
    83,                                    # PB
    84,                                    # RN
    85, 88,                                # CE
    86, 89,                                # PI
    91, 93, 94,                            # PA
    92, 97,                                # AM
    95,                                    # RR
    96,                                    # AP
    98, 99,                                # MA
}


def limpar_numero(raw) -> dict | None:
    """
    Limpa um numero de telefone de QUALQUER formato brasileiro.

    Exemplos:
        (92) 9 8152-3468   -> {numero: "92981523468", numero_discar: "981523468"}
        +55 92 981523468   -> {numero: "92981523468", numero_discar: "981523468"}
        +55 092 98152-3468 -> {numero: "92981523468", numero_discar: "981523468"}
        55 92 981523468    -> {numero: "92981523468", numero_discar: "981523468"}
        092 981523468      -> {numero: "92981523468", numero_discar: "981523468"}
        92981523468        -> {numero: "92981523468", numero_discar: "981523468"}
        981523468          -> {numero: "981523468",   numero_discar: "981523468"}
        9815-23468         -> {numero: "981523468",   numero_discar: "981523468"}
        +5592981523468     -> {numero: "92981523468", numero_discar: "981523468"}

    Returns:
        dict com 'numero' (com DDD se disponivel) e 'numero_discar' (sem DDD),
        ou None se nao for um telefone valido.
    """
    if raw is None:
        return None

    texto = str(raw).strip()
    if not texto:
        return None

    # Passo 1: remover tudo que nao e digito
    digitos = re.sub(r"[^\d]", "", texto)

    if not digitos:
        return None

    # Passo 2: remover codigo do pais 55 (se tiver 12+ digitos e comecar com 55)
    if len(digitos) >= 12 and digitos[:2] == "55":
        digitos = digitos[2:]

    # Passo 3: remover zero inicial (formato antigo de DDD: 092 -> 92)
    if len(digitos) >= 10 and digitos[0] == "0":
        digitos = digitos[1:]

    # Agora digitos deve ter 8-11 caracteres para ser valido
    if len(digitos) < 8 or len(digitos) > 13:
        return None

    # Passo 4: separar DDD e numero local
    numero_completo = digitos
    numero_discar = digitos

    if len(digitos) == 11:
        # DDD (2) + celular (9 digitos comecando com 9)
        ddd_candidato = int(digitos[:2])
        if ddd_candidato in DDDS_BRASIL:
            numero_discar = digitos[2:]  # 9 digitos
        # Mesmo que nao seja DDD valido, aceita como numero
    elif len(digitos) == 10:
        # DDD (2) + fixo (8 digitos)
        ddd_candidato = int(digitos[:2])
        if ddd_candidato in DDDS_BRASIL:
            numero_discar = digitos[2:]  # 8 digitos

    # Validacao final: numero_discar deve ter 8 ou 9 digitos
    if len(numero_discar) < 8 or len(numero_discar) > 9:
        # Pode ser um numero com DDD nao reconhecido mas ainda valido
        # Aceitar se o total de digitos esta entre 8-13
        if 8 <= len(digitos) <= 13:
            return {
                "numero": numero_completo,
                "numero_discar": numero_discar,
            }
        return None

    return {
        "numero": numero_completo,
        "numero_discar": numero_discar,
    }


def _parece_telefone(valor) -> bool:
    """Verifica se um valor parece ser um numero de telefone."""
    resultado = limpar_numero(valor)
    return resultado is not None


def _e_texto_nome(valor) -> bool:
    """Verifica se um valor parece ser um nome (texto, nao numero)."""
    if valor is None:
        return False
    texto = str(valor).strip()
    if not texto:
        return False
    # Se tem mais letras que digitos, e um nome
    letras = sum(1 for c in texto if c.isalpha())
    digitos_count = sum(1 for c in texto if c.isdigit())
    return letras > digitos_count and letras >= 2


def _detectar_delimitador_csv(caminho: str) -> str:
    """Detecta o delimitador de um arquivo CSV."""
    try:
        with open(caminho, "r", encoding="utf-8", errors="replace") as f:
            amostra = f.read(8192)
        try:
            dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
            return dialeto.delimiter
        except csv.Error:
            # Contar ocorrencias de cada delimitador
            contagem = {d: amostra.count(d) for d in [";", ",", "\t", "|"]}
            melhor = max(contagem, key=contagem.get)
            if contagem[melhor] > 0:
                return melhor
            return ","
    except Exception:
        return ","


def _ler_txt(caminho: str) -> list[dict]:
    """Le numeros de um arquivo TXT (um por linha)."""
    numeros = []
    vistos = set()

    try:
        with open(caminho, "r", encoding="utf-8", errors="replace") as f:
            for linha_num, linha in enumerate(f, 1):
                linha = linha.strip()
                if not linha:
                    continue
                resultado = limpar_numero(linha)
                if resultado and resultado["numero"] not in vistos:
                    vistos.add(resultado["numero"])
                    numeros.append({
                        "linha": linha_num,
                        "numero": resultado["numero"],
                        "numero_discar": resultado["numero_discar"],
                        "nome": "",
                    })
    except Exception as e:
        logger.error(f"Erro lendo TXT {caminho}: {e}")

    return numeros


def _ler_csv(caminho: str) -> list[dict]:
    """Le numeros de um arquivo CSV, detectando delimitador automaticamente."""
    delimitador = _detectar_delimitador_csv(caminho)
    logger.info(f"CSV detectado com delimitador: {repr(delimitador)}")

    numeros = []
    vistos = set()

    try:
        with open(caminho, "r", encoding="utf-8", errors="replace") as f:
            leitor = csv.reader(f, delimiter=delimitador)
            linhas = list(leitor)
    except Exception as e:
        logger.error(f"Erro lendo CSV {caminho}: {e}")
        return []

    if not linhas:
        return []

    # Analisar todas as colunas para encontrar quais tem telefones
    num_colunas = max(len(row) for row in linhas) if linhas else 0
    colunas_telefone = _detectar_colunas_telefone_grid(linhas, num_colunas)

    if not colunas_telefone:
        logger.warning("Nenhuma coluna de telefone detectada no CSV.")
        return []

    for col_tel in colunas_telefone:
        col_nome = _detectar_coluna_nome_grid(linhas, num_colunas, col_tel)

        # Determinar se a primeira linha e cabecalho
        primeira_linha = 0
        if linhas and col_tel < len(linhas[0]):
            primeiro_val = linhas[0][col_tel]
            if not _parece_telefone(primeiro_val) and _e_texto_nome(primeiro_val):
                primeira_linha = 1

        for idx_linha in range(primeira_linha, len(linhas)):
            row = linhas[idx_linha]
            if col_tel >= len(row):
                continue
            celula = row[col_tel]
            resultado = limpar_numero(celula)
            if resultado and resultado["numero"] not in vistos:
                vistos.add(resultado["numero"])
                nome = ""
                if col_nome is not None and col_nome < len(row):
                    val_nome = row[col_nome]
                    if _e_texto_nome(val_nome):
                        nome = str(val_nome).strip()
                numeros.append({
                    "linha": idx_linha + 1,
                    "numero": resultado["numero"],
                    "numero_discar": resultado["numero_discar"],
                    "nome": nome,
                })

    return numeros


def _detectar_colunas_telefone_grid(linhas: list, num_colunas: int) -> list[int]:
    """
    Detecta quais colunas contem telefones em um grid (lista de listas).
    Retorna lista de indices de colunas com telefones.
    """
    if not linhas or num_colunas == 0:
        return []

    scores = {}
    total_linhas = len(linhas)
    # Verificar ate 200 linhas para performance
    linhas_check = min(total_linhas, 200)

    for col_idx in range(num_colunas):
        score = 0
        total_com_valor = 0
        for row_idx in range(linhas_check):
            if col_idx < len(linhas[row_idx]):
                val = linhas[row_idx][col_idx]
                if val is not None and str(val).strip():
                    total_com_valor += 1
                    if _parece_telefone(val):
                        score += 1
        if score >= 2 or (score >= 1 and total_com_valor <= 3):
            # Pelo menos 30% das celulas com valor devem parecer telefone
            if total_com_valor > 0 and (score / total_com_valor) >= 0.3:
                scores[col_idx] = score

    # Ordenar por score decrescente
    return sorted(scores.keys(), key=lambda c: -scores[c])


def _detectar_coluna_nome_grid(linhas: list, num_colunas: int, col_tel: int) -> int | None:
    """Detecta coluna de nome adjacente a coluna de telefone em um grid."""
    # Verificar colunas adjacentes (esquerda e direita)
    candidatas = []
    if col_tel > 0:
        candidatas.append(col_tel - 1)
    if col_tel < num_colunas - 1:
        candidatas.append(col_tel + 1)

    for col_candidata in candidatas:
        score_nome = 0
        total_check = 0
        linhas_check = min(len(linhas), 50)
        for row_idx in range(linhas_check):
            if col_candidata < len(linhas[row_idx]):
                val = linhas[row_idx][col_candidata]
                if val is not None and str(val).strip():
                    total_check += 1
                    if _e_texto_nome(val):
                        score_nome += 1
        if total_check > 0 and (score_nome / total_check) >= 0.5:
            return col_candidata

    return None


def _ler_excel(caminho: str, coluna: str = None, aba: str = None) -> list[dict]:
    """Le numeros de um arquivo Excel (.xlsx/.xls)."""
    try:
        wb = load_workbook(caminho, data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"Erro abrindo Excel {caminho}: {e}")
        return []

    ws = wb[aba] if aba and aba in wb.sheetnames else wb.active

    logger.info(f"Lendo planilha: {caminho} (aba: {ws.title})")

    # Ler todas as celulas em um grid (lista de listas)
    linhas = []
    try:
        for row in ws.iter_rows():
            linha = []
            for cell in row:
                try:
                    val = cell.value
                except Exception:
                    val = None
                linha.append(val)
            linhas.append(linha)
    except Exception as e:
        logger.error(f"Erro lendo celulas: {e}")
        wb.close()
        return []

    wb.close()

    if not linhas:
        return []

    num_colunas = max(len(row) for row in linhas)
    logger.info(f"Dimensoes: {len(linhas)} linhas x {num_colunas} colunas")

    numeros = []
    vistos = set()

    # Se o usuario especificou uma coluna, tentar encontra-la
    if coluna:
        col_idx = _encontrar_coluna_por_nome(linhas, num_colunas, coluna)
        if col_idx is not None:
            colunas_telefone = [col_idx]
        else:
            logger.warning(f"Coluna '{coluna}' nao encontrada. Detectando automaticamente...")
            colunas_telefone = _detectar_colunas_telefone_grid(linhas, num_colunas)
    else:
        colunas_telefone = _detectar_colunas_telefone_grid(linhas, num_colunas)

    if not colunas_telefone:
        # Fallback: escanear TODAS as celulas individualmente
        logger.info("Nenhuma coluna predominante. Escaneando todas as celulas...")
        for row_idx, row in enumerate(linhas):
            for col_idx, val in enumerate(row):
                resultado = limpar_numero(val)
                if resultado and resultado["numero"] not in vistos:
                    vistos.add(resultado["numero"])
                    numeros.append({
                        "linha": row_idx + 1,
                        "numero": resultado["numero"],
                        "numero_discar": resultado["numero_discar"],
                        "nome": "",
                    })
        return numeros

    # Para cada coluna de telefone detectada, extrair numeros
    for col_tel in colunas_telefone:
        col_nome = _detectar_coluna_nome_grid(linhas, num_colunas, col_tel)
        logger.info(f"Coluna telefone: {col_tel + 1}" +
                     (f", coluna nome: {col_nome + 1}" if col_nome is not None else ""))

        for row_idx, row in enumerate(linhas):
            if col_tel >= len(row):
                continue
            val = row[col_tel]
            resultado = limpar_numero(val)
            if resultado and resultado["numero"] not in vistos:
                vistos.add(resultado["numero"])
                nome = ""
                if col_nome is not None and col_nome < len(row):
                    val_nome = row[col_nome]
                    if _e_texto_nome(val_nome):
                        nome = str(val_nome).strip()
                numeros.append({
                    "linha": row_idx + 1,
                    "numero": resultado["numero"],
                    "numero_discar": resultado["numero_discar"],
                    "nome": nome,
                })

    return numeros


def _encontrar_coluna_por_nome(linhas: list, num_colunas: int, nome_coluna: str) -> int | None:
    """Encontra o indice de uma coluna pelo nome do cabecalho."""
    nome_lower = nome_coluna.strip().lower()
    # Verificar as primeiras 5 linhas como possiveis cabecalhos
    for row_idx in range(min(5, len(linhas))):
        for col_idx in range(min(num_colunas, len(linhas[row_idx]))):
            val = linhas[row_idx][col_idx]
            if val is not None and str(val).strip().lower() == nome_lower:
                return col_idx
    return None


def ler_numeros(caminho_excel: str, coluna_telefone: str = None, aba: str = None) -> list[dict]:
    """
    Le numeros de telefone de qualquer arquivo (Excel, CSV, TXT).
    Detecta automaticamente colunas, formatos e layouts.

    Args:
        caminho_excel: Caminho do arquivo (.xlsx, .xls, .csv, .txt)
        coluna_telefone: Nome da coluna com os telefones (auto-detecta se None)
        aba: Nome da aba do Excel (usa a primeira se None)

    Returns:
        Lista de dicts: [{"linha": 2, "numero": "92981523468", "numero_discar": "981523468", "nome": "Joao"}, ...]
    """
    if not os.path.exists(caminho_excel):
        raise FileNotFoundError(f"Arquivo nao encontrado: {caminho_excel}")

    ext = os.path.splitext(caminho_excel)[1].lower()
    logger.info(f"Lendo arquivo: {caminho_excel} (formato: {ext})")

    if ext in (".xlsx", ".xls"):
        numeros = _ler_excel(caminho_excel, coluna_telefone, aba)
    elif ext == ".csv":
        numeros = _ler_csv(caminho_excel)
    elif ext == ".txt":
        numeros = _ler_txt(caminho_excel)
    else:
        # Tentar como texto puro
        logger.info(f"Extensao {ext} nao reconhecida, tentando como texto...")
        numeros = _ler_txt(caminho_excel)

    logger.info(f"Total de numeros encontrados: {len(numeros)}")

    # Garantir que todos os dicts tenham as chaves esperadas
    for n in numeros:
        n.setdefault("nome", "")
        n.setdefault("numero_discar", n.get("numero", ""))

    return numeros


def salvar_resultados(resultados: list[dict], caminho_saida: str = None):
    """
    Salva os resultados em uma nova planilha Excel formatada.
    """
    if not caminho_saida:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs("resultados", exist_ok=True)
        caminho_saida = f"resultados/validacao_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Cabecalho
    headers = ["#", "Nome", "Telefone", "Resultado", "Confianca", "Transcricao", "Padrao Detectado", "Horario"]
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Dados
    for i, res in enumerate(resultados, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=res.get("nome", "")).border = border
        ws.cell(row=row, column=3, value=res.get("numero", "")).border = border

        # Resultado com cor
        cell_resultado = ws.cell(row=row, column=4, value=res.get("descricao", ""))
        cor = CORES.get(res.get("codigo", ""), "FFFFFF")
        cell_resultado.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        if cor in ["FF0000", "A020F0"]:
            cell_resultado.font = Font(color="FFFFFF", bold=True)
        else:
            cell_resultado.font = Font(bold=True)
        cell_resultado.border = border

        ws.cell(row=row, column=5, value=f"{res.get('confianca', 0):.0%}").border = border
        ws.cell(row=row, column=6, value=res.get("transcricao", "")).border = border
        ws.cell(row=row, column=7, value=res.get("padrao_encontrado", "")).border = border
        ws.cell(row=row, column=8, value=res.get("horario", "")).border = border

    # Ajustar larguras
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 20

    # Resumo na segunda aba
    ws2 = wb.create_sheet("Resumo")
    ws2.cell(row=1, column=1, value="Resumo da Validacao").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws2.cell(row=3, column=1, value=f"Total de numeros: {len(resultados)}")

    # Contagem por categoria
    contagem = {}
    for res in resultados:
        cat = res.get("descricao", "Desconhecido")
        contagem[cat] = contagem.get(cat, 0) + 1

    row = 5
    ws2.cell(row=row, column=1, value="Categoria").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Quantidade").font = Font(bold=True)
    ws2.cell(row=row, column=3, value="Percentual").font = Font(bold=True)

    for cat, qtd in sorted(contagem.items(), key=lambda x: -x[1]):
        row += 1
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=qtd)
        if len(resultados) > 0:
            ws2.cell(row=row, column=3, value=f"{qtd/len(resultados)*100:.1f}%")
        else:
            ws2.cell(row=row, column=3, value="0%")

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 12

    wb.save(caminho_saida)
    logger.info(f"Resultados salvos: {caminho_saida}")
    return caminho_saida
